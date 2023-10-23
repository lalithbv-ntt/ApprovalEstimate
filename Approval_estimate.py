from datetime import datetime, timedelta
import pandas as pd
from openpyxl import load_workbook
import streamlit as st
import base64

import warnings

# Disable warnings globally
warnings.filterwarnings("ignore")

# Set Streamlit page configuration
st.set_page_config(
    layout="wide",  # Use the full width of the page
    initial_sidebar_state="auto",  # Automatically show/hide the sidebar based on content
    page_title="Approval time estimate",  # Title of the app
    page_icon=":calendar:",  # Icon for the app
)

st.header(":calendar: Approval time estimate")
st.write("***")

with st.sidebar:
    submit_date = st.date_input("Submit date", min_value = datetime.now(), format="DD/MM/YYYY")
    gp = st.number_input("GP%", min_value=0.0, step=0.5)
    revenue_size = st.number_input("Revenue", min_value=0, step=50000)
    submit_button = st.button("Submit", key="submit_button")

#To exclude weekends
def subtract_weekday(day, num_days):
    while num_days > 0:
        day -= timedelta(days=1)
        if day.weekday() < 5:  # Monday to Friday (0-4 are weekdays)
            num_days -= 1
    return day


if submit_button:
    approvals_required_all = ['Practice Director', 'Domain Director', 'CDA', 'VP Digital / Sector Lead', 'VP Consulting', 'CEO / CFO']

    if revenue_size <= 50000 and gp >=34:
        approvals_required = approvals_required_all[:3]
        approval_status = ['Review', 'Approve', 'Review']
        condition = 'revenue_size <= 50000 and gp >=34'
    elif revenue_size <= 50000 and gp <34:
        approvals_required = approvals_required_all[:4]
        approval_status = ['Review', 'Review', 'Review', 'Approve']
        condition = 'revenue_size <= 50000 and gp <34'
    elif revenue_size >= 700000 and revenue_size < 22500000 and gp < 22:
        approvals_required = approvals_required_all
        approval_status = ['Review', 'Review', 'Review', 'Review', 'Review', 'Approve'] 
        condition = 'revenue_size >= 700000 and revenue_size < 22500000 and gp < 22'
    elif revenue_size > 50000 and revenue_size < 1000000:
        approvals_required = approvals_required_all[:4]
        approval_status = ['Review', 'Review', 'Review', 'Approve']
        condition = 'revenue_size > 50000 and revenue_size < 1000000'
    elif revenue_size >= 1000000 and revenue_size < 22500000:
        approvals_required = approvals_required_all[:5]
        approval_status = ['Review', 'Review', 'Review', 'Review', 'Approve']
        condition = 'revenue_size >= 1000000 and revenue_size < 22500000'
    elif revenue_size >= 22500000:
        approvals_required = approvals_required_all
        approval_status = ['Review', 'Review', 'Review', 'Review', 'Review', 'Approve'] 
        condition = 'revenue_size >= 22500000'

    approval_estimate = {}

    last_approval_date = subtract_weekday(submit_date, 2)

    approval_estimate[approvals_required[-1]] = [last_approval_date.strftime("%d/%m/%Y"), approval_status[-1]]

    date_estimate_others = last_approval_date

    for i, j in enumerate(approvals_required[:-1][::-1]):
        date_estimate_others = subtract_weekday(date_estimate_others, 1)
        approval_estimate[j] = [date_estimate_others.strftime("%d/%m/%Y"), approval_status[:-1][::-1][i]]

    df_approval_estimates = pd.DataFrame(list(approval_estimate.items()), columns=['Approver', 'Temp'])

    df_approval_estimates[['Approve by date', 'Approval Status']] = df_approval_estimates['Temp'].apply(pd.Series)

    df_approval_estimates = df_approval_estimates.drop('Temp', axis=1)

    df_pivot = df_approval_estimates.pivot(index='Approver', columns='Approval Status', values='Approve by date').reset_index()

    df_pivot.columns.name = None

    df_approval_estimates = df_pivot[['Approver', 'Review', 'Approve']]

    df_approval_estimates['Approver'] = pd.Categorical(df_approval_estimates['Approver'], categories=approvals_required_all, ordered=True)

    df_approval_estimates = df_approval_estimates.sort_values(by='Approver')

    df_approval_estimates.reset_index(inplace=True, drop = True)

    if "gp" in condition:
        st.write(f"The query satisfies the condition: **{condition}**. GP% was used to calculate Review/Approve dates.")
    else:
        st.write(f"The query satisfies the condition: **{condition}**. GP% was not required to calculate Review/Approve dates.")

    image = st.image("Approval Matrix.png", use_column_width=True)

    st.table(df_approval_estimates)

    # Load the Excel file and the specific sheet
    excel_file = "Bid Timeline sample.xlsx"
    sheet_name = "Bid milestones"
    #df = pd.read_excel(excel_file, sheet_name=sheet_name, engine="openpyxl")

    # Open the Excel file using openpyxl
    wb = load_workbook(excel_file)

    # Select the desired sheet
    sheet = wb[sheet_name]

    start_row =22
    end_row = 27
    start_col = 4  
    end_col = 5  

    # Clear the cells in the specified range
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.value = None

    wb.save(excel_file)

    data = [tuple(row) for row in df_approval_estimates[['Review', 'Approve']].to_records(index=False)]

    start_row = 22
    start_col = 4

    for row_idx, row_data in enumerate(data):
        for col_idx, value in enumerate(row_data):
            cell = sheet.cell(row=start_row + row_idx, column=start_col + col_idx)
            cell.value = value

    wb.save(excel_file)

    Revenue_cell_name = "C1"
    GP_cell_name = "C2"
    Submit_date_cell_name = "C3"

    Revenue_cell = sheet[Revenue_cell_name]
    Revenue_cell.value = revenue_size

    GP_cell = sheet[GP_cell_name]
    GP_cell.value = gp/100

    Submit_date_cell = sheet[Submit_date_cell_name]
    Submit_date_cell.value = submit_date.strftime("%d/%m/%Y")

    wb.save(excel_file)


    # with open(excel_file, "rb") as f:
    #     st.download_button("Download Excel File", f.read(), excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # Read the Excel file and encode it as base64
    with open(excel_file, "rb") as file:
        encoded_data = base64.b64encode(file.read()).decode()

    # Create the download link for the Excel file
    #download_link = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{encoded_data}" download={excel_file}>Download Excel File</a>'

    download_link = f"data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{encoded_data}"
    #st.write(download_link, unsafe_allow_html=True)

    # Display a download button
    st.markdown(
        f'<a href="{download_link}" download={excel_file}>'
        '<button style="padding: 10px 20px; background-color: #008CBA; color: #ffffff; border: none; border-radius: 5px; cursor: pointer;">Download Excel File</button>'
        '</a>',
        unsafe_allow_html=True
    )
